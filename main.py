import openpyxl

entrada = openpyxl.load_workbook('xcl_file\Teste_Entradas.xlsx', data_only=True)

#selecting sheet
cclops = entrada['CCLops']
ccprod = entrada['CCProd']
ccproj = entrada['CCProj']
resumo = entrada['Resumo']

#list of columns
cols_cclops = []
cols_ccprod = []
cols_ccproj = []

def iterate_sheets(sheet, sheet_list, row_min, row_max, col_min=None, col_max=None):
    '''
    iterating each column of each sheet(linha 2)
    row = int
    sheet = celll
    '''
    for rows in sheet.iter_rows(min_row=row_min, max_row=row_max):
        for i in range(col_min, col_max):
            sheet_list.append('R$          {:,.2f}'.format(rows[i].value))  # NOTE: formatando

def iterate_sheets_without_format(sheet, sheet_list, row_min, row_max, col_min=None, col_max=None):
    '''
    iteraing each column of each sheet(row 2) (without formatting)
    row = int
    sheet = cell
    '''
    for rows in sheet.iter_rows(min_row=row_min, max_row=row_max):
        for i in range(col_min, col_max):
            sheet_list.append(rows[i].value)

def fill_cell(ws, cell, value):
    '''
    inserting in each cell
    Parametros: ws = sheet, ex = resume
    '''
    ws[f'{cell}'.upper()] = f'{value}'

iterate_sheets(cclops, cols_cclops, 2, 2, 2, 14)
iterate_sheets(ccprod, cols_ccprod, 2, 2, 2, 14)
iterate_sheets(ccproj, cols_ccproj, 2, 2, 2, 14)

''' Loop to go through the cells and their respective values
create a list of the cells you want to populate and their data'''
cells = [('C6', cols_cclops), ('D6', cols_cclops[1:]), ('C7', cols_ccprod), ('D7', cols_ccprod[1:]), ('C8', cols_ccproj), ('D8', cols_ccproj[1:])]

#loop through this list and fill each cell with its corresponding data
for cell, data in cells:
    for i, value in enumerate(data):
        fill_cell(resumo, chr(ord(cell[0])+i) + cell[1:], value)
  
def total_jan_fev(mes, ops_col, prod_col, proj_col):
    
    ops = entrada['CCLops'].cell(2, ops_col).value
    prod = entrada['CCProd'].cell(2, prod_col).value
    proj = entrada['CCProj'].cell(2, proj_col).value
    
    total = ops + prod + proj
    
    total_formatado = 'R$ {:,.2f}'.format(total)
    
    resumo.cell(4, mes).value = total_formatado
    
total_jan_fev(3, 3, 3, 5)  # Janeiro
total_jan_fev(4, 4, 4, 6)  # Fevereiro

'''
This code defines a list of tuples for categories and a list of tuples for months. 
It then uses a loop to iterate through the months and categories and calculate each 
month's total using a sum() function and a generator expression. Finally, 
it formats the total as a string and writes it to the corresponding cell in the summary worksheet.
'''    

# Define a list of tuples with the sheet names and column indices for each category
categories = [('CCLops', 5), ('CCProd', 3), ('CCProj', 5)]

# Define a list of month names and corresponding column indices
months = [('Mar', 5), ('Apr', 6), ('May', 7), ('Jun', 8), ('Jul', 9), ('Aug', 10), ('Sep', 11), ('Oct', 12), ('Nov', 13)]

# Iterate over the months and categories to calculate the total for each month
for month, col in months:
    total = sum(entrada[cat].cell(2, col).value for cat, col in categories)
    total = 'R$ {:,.2f}'.format(total)
    resumo.cell(4, col).value = total

entrada.save('entradas.xlsx')
