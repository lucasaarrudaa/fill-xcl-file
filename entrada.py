import openpyxl

entrada = openpyxl.load_workbook('Teste_Entradas.xlsx', data_only=True)

# selecting sheet
cclops = entrada['CCLops']
ccprod = entrada['CCProd']
ccproj = entrada['CCProj']
resumo = entrada['Resumo']

# list of columns
cols_cclops = []
cols_ccprod = []
cols_ccproj = []

def iterate_sheets(sheet, sheet_list, row_min, row_max, col_min=None, col_max=None):
    '''
    iterating each column of each sheet(linha 2)
    row = int
    sheet = celll
    '''
    for rows in sheet.iter_rows(min_row=row_min, max_row=row_max):
        for i in range(col_min, col_max):
            sheet_list.append('R$          {:,.2f}'.format(rows[i].value))  # NOTE: formatando

def iterate_sheets_without_format(sheet, sheet_list, row_min, row_max, col_min=None, col_max=None):
    '''
    iteraing each column of each sheet(row 2) (without formatting)
    row = int
    sheet = cell
    '''
    for rows in sheet.iter_rows(min_row=row_min, max_row=row_max):
        for i in range(col_min, col_max):
            sheet_list.append(rows[i].value)

def fill_cell(ws, cell, value):
    '''
    inserting in each cell
    Parametros: ws = sheet, ex = resume
    '''
    ws[f'{cell}'.upper()] = f'{value}'


iterate_sheets(cclops, cols_cclops, 2, 2, 2, 14)
iterate_sheets(ccprod, cols_ccprod, 2, 2, 2, 14)
iterate_sheets(ccproj, cols_ccproj, 2, 2, 2, 14)

print(cols_cclops)



fill_cell(resumo, 'C6', cols_cclops[0])
fill_cell(resumo, 'D6', cols_cclops[1])
fill_cell(resumo, 'E6', cols_cclops[2])
fill_cell(resumo, 'F6', cols_cclops[3])
fill_cell(resumo, 'G6', cols_cclops[4])
fill_cell(resumo, 'H6', cols_cclops[5])
fill_cell(resumo, 'I6', cols_cclops[6])
fill_cell(resumo, 'J6', cols_cclops[7])
fill_cell(resumo, 'K6', cols_cclops[8])
fill_cell(resumo, 'L6', cols_cclops[9])
fill_cell(resumo, 'M6', cols_cclops[10])

fill_cell(resumo, 'C7', cols_ccprod[0])
fill_cell(resumo, 'D7', cols_ccprod[1])
fill_cell(resumo, 'E7', cols_ccprod[2])
fill_cell(resumo, 'F7', cols_ccprod[3])
fill_cell(resumo, 'G7', cols_ccprod[4])
fill_cell(resumo, 'H7', cols_ccprod[5])
fill_cell(resumo, 'I7', cols_ccprod[6])
fill_cell(resumo, 'J7', cols_ccprod[7])
fill_cell(resumo, 'K7', cols_ccprod[8])
fill_cell(resumo, 'L7', cols_ccprod[9])
fill_cell(resumo, 'M7', cols_ccprod[10])

fill_cell(resumo, 'C8', cols_ccproj[0])
fill_cell(resumo, 'D8', cols_ccproj[1])
fill_cell(resumo, 'E8', cols_ccproj[2])
fill_cell(resumo, 'F8', cols_ccproj[3])
fill_cell(resumo, 'G8', cols_ccproj[4])
fill_cell(resumo, 'H8', cols_ccproj[5])
fill_cell(resumo, 'I8', cols_ccproj[6])
fill_cell(resumo, 'J8', cols_ccproj[7])
fill_cell(resumo, 'K8', cols_ccproj[8])
fill_cell(resumo, 'L8', cols_ccproj[9])
fill_cell(resumo, 'M8', cols_ccproj[10])
  
resumo[C4:] 

    
total_jan()
total_fev()
total_mar()
total_apr()
total_may()
total_jun()
total_jul()
total_aug()
total_sep()
total_oct()
total_nov()
    
# entrada.save('tst.xlsx')
