import openpyxl

entrada = openpyxl.load_workbook('Teste_Entradas.xlsx', data_only=True)

# selecting sheet
cclops = entrada['CCLops']
ccprod = entrada['CCProd']
ccproj = entrada['CCProj']
resumo = entrada['Resumo']

# list of columns
cols_cclops = []
cols_ccprod = []
cols_ccproj = []

def brl(value):
    '''
    Converts to BRL format
    '''
    return 'R$          {:,.2f}'.format(value)

def iterate_sheets(sheet, sheet_list, row_min, row_max, col_min=None, col_max=None):
    '''
    iterating each column of each sheet(linha 2)
    row = int
    sheet = celll
    '''
    for rows in sheet.iter_rows(min_row=row_min, max_row=row_max):
        for i in range(col_min, col_max):
            sheet_list.append('R$          {:,.2f}'.format(rows[i].value))  # NOTE: formatando

def iterate_sheets_without_format(sheet, sheet_list, row_min, row_max, col_min=None, col_max=None):
    '''
    iteraing each column of each sheet(row 2) (without formatting)
    row = int
    sheet = cell
    '''
    for rows in sheet.iter_rows(min_row=row_min, max_row=row_max):
        for i in range(col_min, col_max):
            sheet_list.append(rows[i].value)

def fill_cell(ws, cell, value):
    '''
    inserting in each cell
    Parametros: ws = sheet, ex = resume
    '''
    ws[f'{cell}'.upper()] = f'{value}'

def get_values(ws, by, to):
    '''
    generate values in range
    Parameters:
            ws: var of sheet
            by: coordinate from
            to: coordinate to
    Returns:
            list like: ['1', '2', '3']
    '''
    vals = []
    wsheet = ws[f'{by}'.upper():f'{to}'.upper()]
    for r in wsheet:
        for x in r:
            vals.append(x.value)
    return vals

def get_cords(ws, by, to):
    '''
    Usage: generate cords in range
    
    EX: get_cords(sheet, 'c2', 'e2') 
    
    Parameters:
            ws: var of sheet
            by: coordinate from
            to: coordinate to
    Returns:
            list like: ['C2', 'D2', 'E2']
    '''

    cords = []
    wsheet = ws[f'{by}'.upper():f'{to}'.upper()]
    for r in wsheet:
        for x in r:
            cords.append(x.coordinate)
    return cords

#columns of sheets to be used to fill 'resumo'
# cords_cclops = get_cords(cclops, 'c2', 'n2')
# cords_prod = get_cords(ccprod, 'e2', 'n2')
# cords_proj = get_cords(ccproj,'c2', 'n2')

def insert_values(ws_1, cords_1_from, cords_1_to, ws_2, cords_2_from, cords_2_to):
    '''
    Insere valores de duas sheets em uma planilha do excel
    
    Steps:
    1) gera uma lista de  valores das celulas sheet que você quer recuperar o valor (identificada com o nro 1)
    2) gera uma lista das coordenadas das celulas sheet que você quer inserir o valor (identificada com o nro 2)
    3) Insere na lista final (identificada com o nro 2)
    Parameters:
            ws_1: ws inicial
            cords_1_from: coords da ws inicial (str)
            cords_1_to: coords da ws final (str)
            ws_2: ws final
            cords_2_from: coords da ws inicial (str)
            cords_2_to: coords da ws final (str)
            NOTE: em colunas merged, tem que inserir na linha de superior e esquerda, se houver merge horizontal.
    Returns:
    '''
    # generating list like: ['1', '2', '3']
    ws_from = get_values(ws_1, cords_1_from, cords_1_to)
    
    for n in ws_from: 
        float(n) # convertendo para float
        round(n, 2) # arredondando para 2 casas decimais
        n = brl(n) #NOTE: convertendo para o formato brl
    # generating list like: ['C2', 'D2', 'E2']
    ws_to = get_cords(ws_2, cords_2_from, cords_2_to)
    
    for c, v in zip(ws_to, ws_from):
        fill_cell(ws_2, c, v)

insert_values(cclops, 'c2', 'm2', resumo, 'c6', 'm6' )

def insert_values_formatted(ws_1, cords_1_from, cords_1_to, ws_2, cords_2_from, cords_2_to, decimals):
    '''
    Insere valores de duas sheets em uma planilha do excel
    
    Steps:
    1) gera uma lista de  valores das celulas sheet que você quer recuperar o valor (identificada com o nro 1)
    2) gera uma lista das coordenadas das celulas sheet que você quer inserir o valor (identificada com o nro 2)
    3) Insere na lista final (identificada com o nro 2)
    Parameters:
            ws_1: ws inicial
            cords_1_from: coords da ws inicial (str)
            cords_1_to: coords da ws final (str)
            ws_2: ws final 
            cords_2_from: coords da ws inicial (str)
            cords_2_to: coords da ws final (str)
            decimals: nro de casas decimais (int)
            NOTE: em colunas merged, tem que inserir na linha de superior e esquerda, se houver merge horizontal.
    Returns:
    '''
    # generating list like: ['1', '2', '3']
    ws_from = get_values(ws_1, cords_1_from, cords_1_to)
    
    for n in ws_from: 
        float(n) # convertendo para float
        round(n, 2) # arredondando para 2 casas decimais
    # generating list like: ['C2', 'D2', 'E2']
    ws_to = get_cords(ws_2, cords_2_from, cords_2_to)
    
    for c, v in zip(ws_to, ws_from):
        fill_cell(ws_2, c, v)
        
def insert_values_brl(ws_1, cords_1_from, cords_1_to, ws_2, cords_2_from, cords_2_to, decimals):
    '''
    Insere valores de duas sheets em uma planilha do excel
    
    Steps:
    1) gera uma lista de  valores das celulas sheet que você quer recuperar o valor (identificada com o nro 1)
    2) gera uma lista das coordenadas das celulas sheet que você quer inserir o valor (identificada com o nro 2)
    3) Insere na lista final (identificada com o nro 2)
    Parameters:
            ws_1: ws inicial
            cords_1_from: coords da ws inicial (str)
            cords_1_to: coords da ws final (str)
            ws_2: ws final
            cords_2_from: coords da ws inicial (str)
            cords_2_to: coords da ws final
            decimals: nro de casas decimais (int)
            NOTE: em colunas merged, tem que inserir na linha de superior e esquerda, se houver merge horizontal.
    Returns:
    '''
    # generating list like: ['1', '2', '3']
    ws_from = get_values(ws_1, cords_1_from, cords_1_to)
    
    for n in ws_from: 
        float(n) # convertendo para float
        round(n, 2) # arredondando para 2 casas decimais
    ws_from_real = []
    
    for n in ws_from:
        ws_from_real.append(brl(n))
    # generating list like: ['C2', 'D2', 'E2']
    ws_to = get_cords(ws_2, cords_2_from, cords_2_to)
    
    for c, v in zip(ws_to, ws_from_real):
        fill_cell(ws_2, c, v)
        
        
insert_values_brl(cclops, 'c2', 'm2', resumo, 'c6', 'm6', 2)
insert_values_brl(ccprod, 'e2', 'm2', resumo, 'e7', 'm7', 2)
insert_values_brl(ccproj, 'c2', 'm2', resumo, 'c8', 'm8', 2)
entrada.save('tst.xlsx')